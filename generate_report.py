from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
done_fill = PatternFill("solid", fgColor="C6EFCE")
hardcode_fill = PatternFill("solid", fgColor="FFC7CE")
config_fill = PatternFill("solid", fgColor="FFEB9C")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ===== Sheet 1: 系统功能配置说明 =====
ws1 = wb.active
ws1.title = "系统功能配置说明"

ws1['A1'] = "COCO CLAW 系统功能配置说明"
ws1['A1'].font = Font(bold=True, size=16)
ws1.merge_cells('A1:F1')

ws1['A2'] = "更新时间: 2026-04-07 22:40"
ws1['A2'].font = Font(size=10, italic=True)

headers = ["模块", "功能点", "当前状态", "类型", "配置方式", "备注"]

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# 数据（不使用合并单元格，改用格式化）
data = [
    ["C端前端", "技能商品展示", "✅ 已完成", "可配置", "管理后台 → 技能管理", "上架/下架、价格、名称、描述"],
    ["C端前端", "技能分类筛选", "✅ 已完成", "可配置", "管理后台 → 分类管理", "添加/编辑/删除分类"],
    ["C端前端", "购物车功能", "✅ 已完成", "可配置", "用户自助操作", "用户自行添加/删除商品"],
    ["C端前端", "订单创建", "✅ 已完成", "可配置", "用户自助操作", "下单时可选收货地址、优惠券"],
    ["C端前端", "支付流程", "✅ 已完成", "❌ 写死", "需同事接入支付", "支付宝/微信支付（待开发）"],
    ["C端前端", "个人中心", "✅ 已完成", "❌ 写死", "代码固定", "用户信息、订单入口"],
    ["C端前端", "收货地址", "✅ 已完成", "可配置", "用户自助操作", "用户自行增删改查地址"],
    ["C端前端", "优惠券使用", "✅ 已完成", "可配置", "管理后台 → 优惠券管理", "发放优惠券给用户"],
    ["C端前端", "免费/付费商品", "⏳ 待开发", "❌ 缺失", "需加字段", "建议加 price_type 字段区分"],
    ["C端前端", "价格展示", "✅ 已完成", "可配置", "管理后台设置", "价格自由填写，无上限限制"],
    ["C端前端", "价格校验", "⏳ 待开发", "❌ 缺失", "需加校验", "建议限制最低0.01、最高99999"],
    ["C端前端", "登录/注册", "✅ 已完成", "可配置", "短信服务商配置", "验证码发送（需对接短信API）"],
    ["C端前端", "用户等级", "⏳ 待开发", "❌ 缺失", "需加功能", "VIP/普通用户区分"],
    ["C端前端", "首页Banner", "✅ 已完成", "❌ 写死", "代码固定", "需改源码或开发后台管理"],
    ["C端前端", "热门推荐", "✅ 已完成", "❌ 写死", "代码固定", "推荐商品写死在代码中"],
    ["C端前端", "搜索功能", "✅ 已完成", "可配置", "用户自助", "按技能名称搜索"],
    ["C端前端", "排序功能", "✅ 已完成", "可配置", "用户自助", "按价格/销量/最新排序"],
    
    ["管理后台", "登录认证", "✅ 已完成", "可配置", "修改源码管理员账号", "目前账号密码写死"],
    ["管理后台", "技能管理", "✅ 已完成", "可配置", "管理后台操作", "增删改查、上架下架、编辑价格"],
    ["管理后台", "分类管理", "✅ 已完成", "可配置", "管理后台操作", "增删改查、父子分类、排序"],
    ["管理后台", "订单管理", "✅ 已完成", "可配置", "管理后台操作", "查看订单、处理退款"],
    ["管理后台", "用户管理", "✅ 已完成", "可配置", "管理后台操作", "查看用户、封禁/解封、余额调整"],
    ["管理后台", "优惠券管理", "✅ 已完成", "可配置", "管理后台操作", "创建优惠券、发放给用户"],
    ["管理后台", "数据统计", "✅ 已完成", "❌ 写死", "代码固定", "统计指标写死在Dashboard中"],
    ["管理后台", "权限控制", "✅ 已完成", "可配置", "代码中配置角色", "管理员/运营/客服角色"],
    ["管理后台", "Banner管理", "⏳ 待开发", "❌ 缺失", "需加功能", "目前Banner写死在C端首页"],
    
    ["后端", "数据库配置", "✅ 已完成", "可配置", "application.yml", "修改数据库连接地址"],
    ["后端", "Redis配置", "✅ 已完成", "可配置", "application.yml", "修改Redis连接地址"],
    ["后端", "Nacos注册", "✅ 已完成", "可配置", "application.yml", "修改Nacos服务地址"],
    ["后端", "支付接口", "⏳ 待开发", "❌ 缺失", "同事负责", "支付宝/微信支付接入"],
    ["后端", "短信接口", "⏳ 待开发", "❌ 缺失", "需对接", "验证码发送（阿里云/腾讯云）"],
    ["后端", "JWT密钥", "✅ 已完成", "可配置", "application.yml", "修改Token加密密钥"],
    ["后端", "分页大小", "✅ 已完成", "❌ 写死", "代码固定", "目前每页30条"],
    ["后端", "接口限流", "⏳ 待开发", "❌ 缺失", "需加功能", "防止恶意请求"],
    ["后端", "日志记录", "⏳ 待开发", "❌ 缺失", "需加功能", "操作日志、异常日志"],
]

for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 类型列填充颜色
        if col_idx == 4:
            if "可配置" in str(value):
                cell.fill = config_fill
            elif "❌ 写死" in str(value) or "❌ 缺失" in str(value):
                cell.fill = hardcode_fill
        
        # 模块列加粗
        if col_idx == 1 and row_data[0]:
            cell.font = Font(bold=True)

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 35
ws1.column_dimensions['F'].width = 30

# ===== Sheet 2: 配置项汇总 =====
ws2 = wb.create_sheet("配置项汇总")

ws2['A1'] = "COCO CLAW 可配置项汇总"
ws2['A1'].font = Font(bold=True, size=16)
ws2.merge_cells('A1:D1')

headers2 = ["配置项", "配置位置", "配置方式", "当前状态"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

config_items = [
    ["技能名称", "管理后台 → 技能管理 → 编辑", "输入框自由填写", "✅ 可用"],
    ["技能价格", "管理后台 → 技能管理 → 编辑", "输入框自由填写", "✅ 可用"],
    ["技能上下架", "管理后台 → 技能管理 → 操作", "点击按钮切换", "✅ 可用"],
    ["技能分类", "管理后台 → 技能管理 → 编辑", "下拉选择分类", "✅ 可用"],
    ["技能描述", "管理后台 → 技能管理 → 编辑", "富文本编辑器", "✅ 可用"],
    ["技能图标", "管理后台 → 技能管理 → 编辑", "上传图片", "✅ 可用"],
    ["技能库存", "管理后台 → 技能管理 → 编辑", "输入数字", "✅ 可用"],
    ["免费/付费", "管理后台 → 技能管理 → 编辑", "单选框切换", "⏳ 待开发"],
    ["添加分类", "管理后台 → 分类管理 → 新增", "输入名称、选择图标", "✅ 可用"],
    ["编辑分类", "管理后台 → 分类管理 → 编辑", "修改名称/图标/排序", "✅ 可用"],
    ["删除分类", "管理后台 → 分类管理 → 删除", "点击删除按钮", "✅ 可用"],
    ["分类排序", "管理后台 → 分类管理 → 编辑", "拖拽或输入序号", "✅ 可用"],
    ["创建优惠券", "管理后台 → 优惠券 → 新增", "满减/折扣/无门槛", "✅ 可用"],
    ["发放优惠券", "管理后台 → 优惠券 → 发放", "选择用户或全部", "✅ 可用"],
    ["优惠券有效期", "管理后台 → 优惠券 → 新增", "设置开始/结束日期", "✅ 可用"],
    ["封禁用户", "管理后台 → 用户管理 → 操作", "点击封禁按钮", "✅ 可用"],
    ["解封用户", "管理后台 → 用户管理 → 操作", "点击解封按钮", "✅ 可用"],
    ["调整余额", "管理后台 → 用户管理 → 余额调整", "输入增减值和原因", "✅ 可用"],
    ["发送消息", "管理后台 → 用户管理 → 发送消息", "输入标题和内容", "✅ 可用"],
    ["查看订单", "管理后台 → 订单管理 → 列表", "筛选条件查看", "✅ 可用"],
    ["处理退款", "管理后台 → 订单管理 → 详情", "同意/拒绝退款", "✅ 可用"],
    ["导出订单", "管理后台 → 订单管理 → 导出", "点击导出按钮", "✅ 可用"],
    ["数据库连接", "后端 application.yml", "修改JDBC URL", "✅ 可用"],
    ["Redis连接", "后端 application.yml", "修改Redis地址端口", "✅ 可用"],
    ["Nacos地址", "后端 application.yml", "修改注册中心地址", "✅ 可用"],
    ["JWT密钥", "后端 application.yml", "修改secret字符串", "✅ 可用"],
    ["管理员账号", "管理后台源码", "修改数据库或代码", "⚠️ 需改代码"],
]

for row_idx, row_data in enumerate(config_items, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_idx == 4:
            if "✅ 可用" in value:
                cell.fill = done_fill
            elif "⏳" in value or "⚠️" in value:
                cell.fill = config_fill

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 15

# ===== Sheet 3: 写死项汇总 =====
ws3 = wb.create_sheet("写死项汇总")

ws3['A1'] = "COCO CLAW 写死项汇总（需改代码）"
ws3['A1'].font = Font(bold=True, size=16)
ws3.merge_cells('A1:D1')

headers3 = ["位置", "功能点", "当前状态", "建议"]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

hardcode_items = [
    ["C端首页", "Banner轮播图", "写死3张图", "建议开发后台Banner管理功能"],
    ["C端首页", "热门推荐位", "写死推荐商品", "建议开发手动推荐功能"],
    ["C端首页", "分类入口展示", "固定显示顺序", "建议允许后台配置显示哪些分类"],
    ["C端前端", "价格货币符号", "固定显示¥", "建议后台可配置货币符号"],
    ["C端前端", "页面布局结构", "固定布局", "如需改需重写页面"],
    ["C端前端", "底部导航项", "固定5个入口", "如需增删需改代码"],
    ["C端前端", "登录注册页样式", "固定样式", "如需改需重写页面"],
    ["管理后台", "数据统计指标", "固定的4个指标", "建议允许自定义统计维度"],
    ["管理后台", "左侧菜单项", "固定菜单结构", "如需改需改源码"],
    ["管理后台", "管理员账号", "代码中写死", "建议改成数据库存储"],
    ["后端", "分页大小", "固定30条/页", "建议后台可配置"],
    ["后端", "Token有效期", "固定7天", "建议后台可配置"],
]

for row_idx, row_data in enumerate(hardcode_items, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = hardcode_fill

ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 35

# ===== Sheet 4: 待开发项 =====
ws4 = wb.create_sheet("待开发项")

ws4['A1'] = "COCO CLAW 待开发项"
ws4['A1'].font = Font(bold=True, size=16)
ws4.merge_cells('A1:E1')

headers4 = ["功能", "所属模块", "优先级", "负责人", "说明"]
for col, header in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

todo_items = [
    ["支付接口接入", "后端", "高", "同事", "支付宝/微信支付对接"],
    ["免费/付费区分", "全栈", "高", "可找我开发", "加price_type字段区分"],
    ["价格校验", "后端+前端", "中", "可找我开发", "限制0.01~99999，2位小数"],
    ["Banner管理", "管理后台+C端", "中", "可找我开发", "后台上传Banner图"],
    ["热门推荐后台配置", "管理后台+C端", "中", "可找我开发", "后台手动设置推荐商品"],
    ["管理员账号后台管理", "管理后台", "中", "可找我开发", "从数据库读取，非代码写死"],
    ["短信接口对接", "后端+前端", "中", "同事", "验证码发送（阿里云/腾讯云）"],
    ["用户等级系统", "全栈", "低", "可找我开发", "VIP/普通用户区分"],
    ["操作日志记录", "后端", "低", "同事", "记录管理员操作"],
    ["接口限流", "后端", "低", "同事", "防止恶意请求"],
    ["单元测试", "后端", "低", "同事", "Junit5测试用例"],
    ["Swagger文档", "后端", "低", "同事", "接口文档自动生成"],
]

for row_idx, row_data in enumerate(todo_items, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_idx == 3:
            if value == "高":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
            elif value == "中":
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
            elif value == "低":
                cell.fill = done_fill

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 10
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 30

wb.save('C:/Users/ROG/CodeBuddy/20260404122544/COCO_CLAW_代码开发进度.xlsx')
print("Excel 文件已更新")
